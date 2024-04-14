# retrieves bodyweights, then writes them to the correct row in the target file (specified in params.py).
from collections import UserDict
from datetime import datetime, timedelta
from typing import List, Tuple

import openpyxl

import utilities.params as p
import utilities.utility_functions as uf
from utilities.shared_types import Entry


class RowBodyweightPairings(UserDict):
    # todo: add tests for this
    def __setitem__(self, row: int, bodyweight: str | float | int):
        # disallow duplicate rows, or the updating of values
        assert isinstance(row, int), "Row must be an integer"
        assert row not in self.data.keys(), "This key has already been used"

        # validate bodyweight
        self.validate_bodyweight_text(bodyweight)
        self.data[int(row)] = str(bodyweight)

    @staticmethod
    def validate_bodyweight_text(bodyweight: str):
        bodyweight = str(bodyweight)
        assert len(bodyweight) > 0, "Empty bodyweight received"

        legal_chars = list("0123456789?.")
        invalid_chars = [c for c in bodyweight if c not in legal_chars]
        assert not invalid_chars, \
            f"Provided bodyweight `{bodyweight}` is invalid. The following characters are not accepted`{invalid_chars}`"

        if "?" not in bodyweight:
            try:
                float(bodyweight)
            except ValueError:
                raise ValueError("Provided bodyweight is invalid and cannot be converted to float.")


def return_most_recent_bodyweights(bodyweights: List[str], desired_count: int) -> List[str] | None:
    """
    Return the X elements with the greatest index, stripped of spaces
    :param bodyweights: a list of strings
    :param desired_count: how many items to include in the returned list
    :return: a list of string bodyweights
    """
    if not any(bodyweights):
        return None

    # we need string values because only strings can represent question marks "?" (i.e. absent bodyweights)
    bodyweights = [str(weight) for weight in bodyweights]

    # we can't return a negative count
    assert isinstance(desired_count, int), "Please provide an integer history_length"
    desired_count = max(desired_count, 0)

    try:
        history = bodyweights[-desired_count:]
    except IndexError:
        # history greater than the number of bodyweights in note
        history = bodyweights

    return [entry.replace(' ', '') for entry in history]


def format_bodyweight_history(history: List[str]) -> str:
    """
    Take a string, parenthesize it, and comma-separate its elements. Include a trailing comma and space. If there's no
    history to format, then return the empty string
    :param history: a list of strings
    :return: a formatted string
    """
    if len(history) == 0 or (len(history) == 1 and history[0] == ""):
        return ''
    return f"(" + ", ".join(history) + "), "


def return_depunctuated_bodyweights_text(text,
                                         keep_decimal_separator=False,
                                         keep_spaces=False,
                                         keep_question_marks=False) -> str:
    """
    Given an input string, return a copy with punctuation partially removed.
    :param text: the input string
    :param keep_decimal_separator: if False, then remove "." chars from string
    :param keep_spaces: if False, then remove spaces from string
    :param keep_question_marks: if False, then remove question marks from string
    :return: a string with less punctuation
    """
    txt = text.replace(",", "").replace("(", "").replace(")", "")
    if not keep_decimal_separator:
        txt = txt.replace(".", "")
    if not keep_spaces:
        txt = txt.replace(" ", "")
    if not keep_question_marks:
        txt = txt.replace("?", "")
    return txt


def extract_bodyweights_from_string(raw_string, split_on_parenthesis: bool) -> List[str] | Tuple[List[str], List[str]]:
    """
    Given a validated string, return the list of bodyweights found in that string. If split_on_parentheses,
    then return two lists - one containing values found inside parentheses, and one containing those outside
    :param raw_string: a string containing comma-separated floats, ints, "?", or parentheses etc.
    :param split_on_parenthesis: whether to split on parentheses (typically used for the context window)
    :return: one or two lists, containing bodyweights found in each group
    """

    validate_bodyweight_note_text(raw_string)
    depunc_str = return_depunctuated_bodyweights_text(raw_string,
                                                      keep_decimal_separator=False,
                                                      keep_spaces=False,
                                                      keep_question_marks=True)
    if len(depunc_str.replace(' ', '')) == 0:
        # No bodyweights to return
        if split_on_parenthesis:
            return [], []
        return []

    bodyweights = [val.strip().replace("(", "").replace(")", "")
                   for val in raw_string.split(',')
                   if val.strip()]

    if not split_on_parenthesis:
        # Return all bodyweights in string
        return bodyweights

    # Return 2 lists, split on closing parenthesis
    if ")" in raw_string:
        assert raw_string.count(")") == 1, "Too many ')' found in string"
        raw_string = raw_string.strip("(").replace(",", " ")
        split_1, split_2 = raw_string.split(")")
        context_window_weights = [val for val in split_1.split() if val.strip()]
        uncommitted_weights = [val for val in split_2.split() if val.strip()]
        return context_window_weights, uncommitted_weights
    else:
        return [], bodyweights


def validate_bodyweight_note_text(bw_note_text: str) -> None:
    """
    If the bodyweight note text is not formatted as expected, then raise an exception
    :param bw_note_text: the string text found within the bodyweight note
    """

    text = bw_note_text
    if "(" in text and text.index("(") != 0:
        raise ValueError("If an opening parenthesis is in the bodyweights note, it must be at the beginning")
    if text.count("(") != text.count(")"):
        raise ValueError("Mismatched parentheses in bodyweights note")
    if text.count("(") > 1 or text.count(")") > 1:
        raise ValueError("Too many parentheses found in bodyweights note. Expected no more than 1 pair of "
                         "opening or closing parentheses")
    if "()" in text:
        raise ValueError("Empty parentheses in bodyweights note. If parentheses are provided in the note, then "
                         "they must surround at least 1 bodyweight")

    de_punctuated_text = return_depunctuated_bodyweights_text(text).replace('\n', '')
    if len(de_punctuated_text) > 0 and not de_punctuated_text.isdigit():
        raise ValueError("Incorrectly formatted list of bodyweight provided in the bodyweights note")


def pair_new_bodyweights_with_rows(sheet, bodyweights: List[float | str], start_row: int, max_rows_without_date=10) \
        -> RowBodyweightPairings:
    """
    :param sheet: sheet in xlsx file containing bodyweights and dates
    :param bodyweights: list of bodyweights not yet committed to file
    :param start_row: the row at which the search starts
    :param max_rows_without_date: the number of rows without a valid date, after which an error is raised, if a
    suitable row cannot be found.
    :return: a list of tuples, where tuple[0] is the int row to write to, and tuple[1] the str bodyweight. Empty cells
    are accounted for.
    """

    assert isinstance(start_row, int)
    assert isinstance(max_rows_without_date, int)

    current_row = start_row
    count_empty = 0
    pairings = RowBodyweightPairings()

    # TODO: refactor
    for bw in bodyweights:
        success = False
        while not success:
            try:
                success = uf.convert_string_to_datetime(sheet.cell(row=current_row, column=p.DATE_COLUMN).value)
            except ValueError:
                # skip empty cells in date column (e.g. at end of year), up to max length "max_empty_rows"
                current_row += 1
                count_empty += 1
                if count_empty >= max_rows_without_date:
                    raise RuntimeError(
                        "Failed to pair bodyweights with rows matching those bodyweights' entry dates. Too many date "
                        f"cells in the Excel sheet are missing datetime values (the cutoff is "
                        f"{max_rows_without_date}). Please verify that the date cell column in your Excel sheet "
                        "contains enough dates"
                    )

        # check if bodyweight cell is already written to
        if sheet.cell(row=current_row, column=p.BODYWEIGHT_COLUMN).value is None:
            pairings[current_row] = bw
            current_row += 1
        else:
            raise RuntimeError(f"Bodyweight cannot be written to target row {current_row} - cell already written to!"
                               f"No changes have been made")

    return pairings


def write_to_file(wb, sheet, row_bodyweight_pairings: RowBodyweightPairings) -> None:
    """
    Write bodyweights to file. Expect validation to be done prior.
    :param wb: the workbook to write to
    :param sheet: the target sheet within the workbook
    :param row_bodyweight_pairings: the row and bodyweight pairings. A UserDict with validated entries
    """

    for row, bodyweight in row_bodyweight_pairings.items():
        try:
            # We write as float because otherwise Calc (and perhaps Excel) prepend each value with a "'", to mark it as
            # a string, causing it to be left-aligned. The float conversion avoids that
            sheet.cell(row=row, column=p.BODYWEIGHT_COLUMN).value = float(bodyweight)
        except ValueError:
            # The given bodyweight is probably "?"
            sheet.cell(row=row, column=p.BODYWEIGHT_COLUMN).value = bodyweight

    wb.save(p.TARGET_PATH)


def main():
    uf.validate_target_sheet_params()

    # use preferred retrieval method to retrieve notes
    handler = uf.return_handler()

    wb = openpyxl.load_workbook(p.TARGET_PATH)
    sheet = wb[p.TARGET_SHEET]
    bw_note: Entry = handler.return_bodyweights_note()

    # if this program is run after 5 AM, then expect the note to have been edited today. Else, yesterday.
    today = datetime.now()
    if datetime.now().hour < 5:
        today -= timedelta(days=1)

    if bw_note.edit_timestamp < today.replace(hour=0, minute=0, second=0, microsecond=0):
        print("- You have not edited your bodyweights note today.")
        print("- Please add today's bodyweight to the note. Then run the program again.")
        print("- If you don't remember it, a question mark will be fine.")
        print(f"Note edit timestamp={bw_note.edit_timestamp}, note text=\"{bw_note.text}\"")
        exit()

    start_row = uf.return_first_absent_bodyweight_row(sheet,
                                                      date_column=p.DATE_COLUMN,
                                                      bodyweight_column=p.BODYWEIGHT_COLUMN)
    todays_row = uf.find_row_of_cell_matching_datetime(sheet, today, date_column=p.DATE_COLUMN)

    if start_row == -1:
        raise RuntimeError("Start row not found")
    if todays_row == -1:
        raise RuntimeError("Failed to find the date cell corresponding to today's date in the xlsx file")
    if sheet.cell(todays_row, p.BODYWEIGHT_COLUMN).value:
        print("Today's bodyweight is already written to file. Exiting program")
        exit()

    # Separate the bodyweights that have been committed to file (which are saved in the context window) from those
    # that have not
    _, uncommitted_bodyweights = extract_bodyweights_from_string(raw_string=bw_note.text,
                                                                 split_on_parenthesis=True)

    if len(uncommitted_bodyweights) == 0:
        print("INFO: no bodyweights found in note. There is nothing new to write\nExiting")
        exit()

    # We expect the bodyweights note to contain one bodyweight per missing entry in the Excel file
    num_expected_bodyweights = (todays_row - start_row + 1)
    if num_expected_bodyweights != len(uncommitted_bodyweights):
        raise ValueError(
            f"Number of bodyweights provided in the bodyweights note ({len(uncommitted_bodyweights)}) does not match "
            f"the number of days for which a bodyweight is expected ({num_expected_bodyweights}).\nPlease correct the "
            "note. If you've forgotten a value, then a question mark is a valid substitute for that bodyweight."
        )

    # pair bodyweights with their target rows. Account for empty rows, and raise if anything is amiss.
    row_bodyweight_mapping: RowBodyweightPairings = pair_new_bodyweights_with_rows(sheet=sheet,
                                                                                   bodyweights=uncommitted_bodyweights,
                                                                                   start_row=start_row)

    # prepare history (or "context window") of the most recently committed-to-file bodyweights, to be written to the
    # bodyweight note
    all_bodyweights = extract_bodyweights_from_string(bw_note.text, split_on_parenthesis=False)
    most_recent_bodyweights: List[str] = return_most_recent_bodyweights(bodyweights=all_bodyweights,
                                                                        desired_count=p.HISTORY_LENGTH)
    history: str = format_bodyweight_history(most_recent_bodyweights)

    uf.backup_file_to_dir(source_file_name=p.TARGET_PATH, backup_directory=p.LOCAL_EXCEL_BACKUP_DIR)
    print("Writing bodyweights to file")
    write_to_file(wb, sheet, row_bodyweight_mapping)

    uf.backup_file_to_dir(source_file_name=bw_note.path, backup_directory=p.LOCAL_NOTES_ARCHIVE_DIR)

    # all done. We can replace the bodyweights note
    print("Updating bodyweights note")
    handler.replace_bodyweights_note(new_text=history)
    print("Finished!")


if __name__ == '__main__':
    main()
