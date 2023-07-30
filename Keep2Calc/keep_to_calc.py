import re
import openpyxl
import os
from datetime import datetime
from dateutil.parser import parse
from dataclasses import dataclass
from typing import Dict, List
import gkeepapi.node
import utilities.params as p
import utilities.utility_functions as uf


# todo: add option to have year in note title, and to use that if provided?
# todo: add tests. Make it clearer what each of these functions does.


@dataclass
class ParsedWorkout:
    # this class will hold processed data that isn't ready to write yet (it still needs a target row).

    # title_datetime is the interpreted date time of the note's title
    title_datetime: datetime

    # the formatted workout data
    data: str

    def __repr__(self):
        return f"{self.title_datetime}: {self.data}"


@dataclass
class DataToWrite(ParsedWorkout):
    # this will hold processed data that's validated, ready to write, and paired with a target row
    target_row: "int > 0"

    def __repr__(self):
        return f"<{self.title_datetime}, target_row={self.target_row}>: {self.data}"


def initial_checks(notes_lst: List[gkeepapi.node.Note]) -> None:
    """
    Run basic validation on the passed in object, and on the values set in params.py. Raise on failure.
    :param notes_lst: a list of Note objects
    """
    # TODO: check their types
    for note in notes_lst:
        if not isinstance(note, gkeepapi.node.List) and not isinstance(note, gkeepapi.node.Node):
            raise ValueError(f"Invalid type found in notes list. "
                             f"Expected type gkeepapi objects, but found {type(note)}")
    if not os.path.exists(p.TARGET_PATH):
        raise FileNotFoundError("Target path not found")
    if not uf.target_path_is_xslx(p.TARGET_PATH):
        raise ValueError(f"Target file is not xslx. This program is not intended for use with non-xlsx target files.")
    if not uf.target_sheet_exists(p.TARGET_PATH, p.TARGET_SHEET):
        raise ValueError(f"Error: target sheet `{p.TARGET_SHEET}` not found in file at {p.TARGET_PATH}")


def _load_workout_notes_to_dict(validated_workout_notes: List[gkeepapi.node.Note], regress_future_dates=False) \
        -> Dict[datetime, str]:
    """
    Load the info of workout_notes into a dictionary, where each key is the note's title, and each value is the same
    note's text. Each Note should have a unique date as its title.
    :param validated_workout_notes: a list of workout notes, as Note objects
    :param regress_future_dates: if a note title would evaluate to a future date, then subtract a year from it
    :return: a dictionary
    """

    # we expect workout note titles to be like: "26 January" or "02 March 2022"
    workouts_dict = dict()
    for note in validated_workout_notes:
        # remove fluff from note title
        parsed_title = re.sub(re.compile(r", day \d"), "", note.title)
        parsed_title = re.sub(re.compile("(,)? off day"), "", parsed_title)

        try:
            # convert user-provided date, using their provided year
            datetime_title = uf.convert_string_to_datetime(date_str=parsed_title,
                                                           regress_future_dates=regress_future_dates)
        except ValueError as e:
            raise ValueError("Received note without valid date in title. All workouts must contain a date in their "
                             f"title.\nMore info: the note was last edited at {note.timestamps.edited}, and contains "
                             f"the following text\n{note.text}") from e

        if datetime_title.year == 1900:
            # user did not provide year. We therefore reconvert, using the current year
            parsed_title += f" {datetime.now().year}"
            datetime_title = uf.convert_string_to_datetime(date_str=parsed_title,
                                                           regress_future_dates=regress_future_dates)

        # this shouldn't happen, because the notes should already be validated.
        assert datetime_title not in workouts_dict.keys(), f"Failed to load workout to dict. Key `{datetime_title}`" \
                                                           f" is present multiple times"
        workouts_dict[datetime_title] = note.text

    return workouts_dict


def parse_workout_notes(validated_workout_notes: List[gkeepapi.node.Note]) -> List[ParsedWorkout]:
    """
    Given a list of workout Note objects, clean up and format the workout within each Note, then return it as a list
    of ParsedWorkout objects, each representing one workout in its final format.
    :param validated_workout_notes: a list of Note objects, each representing a workout. Title dates should be unique.
    :return: a list of ParsedWorkout objects
    """

    for note in validated_workout_notes:
        if not uf.str_contains_est_xx_mins_line(note.text):
            raise ValueError("Received a Note object whose text does not contain an est xx mins line. All workout "
                             "Notes are expected to have such a line")

    workout_notes_dict = _load_workout_notes_to_dict(validated_workout_notes, regress_future_dates=True)

    parsed_data_lst = []
    for datetime_title, workout_text in workout_notes_dict.items():
        # strip lines, and drop empty lines and comment lines.
        workout_text = [line.strip() for line in workout_text.split('\n')
                        if line
                        and not (line_is_comment(line) or line.startswith('\n'))]

        # clean up and format one workout
        one_workout_lines = []
        for line in workout_text:
            parsed_line = capitalize_selectively(line)

            # clean-up
            parsed_line = parsed_line.replace(';', '.')
            parsed_line = parsed_line.replace('..', '.')
            parsed_line = parsed_line.replace(' .', '.')
            parsed_line = parsed_line.replace(',,', '.')
            parsed_line = parsed_line.replace('\n', '')
            parsed_line = parsed_line.replace('  ', ' ')

            # the "+" symbol can be used at the beginning of the line in a note, to indicate an "extra" exercise (i.e.
            # one not part of the standard workout). We will not write it to file.
            parsed_line = parsed_line.replace('+ ', '')
            parsed_line = parsed_line.replace('+', '')

            # hard-coded replacement strings. This is personal preference.
            if "home workout" in line.lower():
                # discard subsequent information, e.g. collapsing the following onto the same string:
                # "Home workout, upper body A:", "Home workout, upper body B:", "Home workout, lower body + abs:"
                parsed_line = "Home workout: "
            if "shadowboxing" in line.lower():
                parsed_line = "Shadowboxing: "

            # this strips certain instructions from the line.
            # parsed_line = strip_num_x_nums(parsed_line)

            # trailing semi-colons can happen due to data entry errors
            if parsed_line.endswith(":"):
                parsed_line = parsed_line[:-1]

            parsed_line = parsed_line.rstrip()

            # add that cleaned line to the workout's lines
            one_workout_lines.append(parsed_line)

        # semicolon-separate each exercise
        exercises_str = '; '.join(one_workout_lines)

        # replace the final semicolon with a full stop.
        exercises_str, est_xx_mins_line = exercises_str.rsplit('; ', 1)
        complete_workout_text = exercises_str + ". " + est_xx_mins_line

        # save the formatted workout
        parsed_data_lst.append(ParsedWorkout(title_datetime=datetime_title, data=complete_workout_text))

    return parsed_data_lst


def pair_workouts_with_rows(parsed_workouts: List[ParsedWorkout]) -> List[DataToWrite]:
    """
    Given a list of parsed workouts, pair each workout with a unique row in the target file. That row's date column cell
    value equals the date of the workout's interpreted datetime. Return that data as list of objects containing the
    interpreted date, the workout, and the target row to write to.
    :param parsed_workouts: a list of fully formatted workouts, of type ParsedWorkout
    :return: a list of parsed workouts, paired with suitable row number.
    """

    wb = openpyxl.load_workbook(p.TARGET_PATH)
    sheet = wb[p.TARGET_SHEET]

    # collect errors, for later
    failed_to_find_date_cell = []
    workout_already_written = []
    target_cell_contains_clashing_info = []

    # the list to be returned
    workouts_to_write: List[DataToWrite] = []

    # todo: fix. Ugly locally, and ugly logically. This paragraph shouldn't be necessary
    if not len(parsed_workouts):
        print("No workouts to write")
        exit()

    for workout in parsed_workouts:
        workout_datetime = workout.title_datetime
        row_match = uf.find_row_of_cell_matching_datetime(sheet,
                                                          workout_datetime,
                                                          p.DATE_COLUMN,
                                                          raise_on_failure=False)
        if row_match == -1:
            failed_to_find_date_cell.append(workout)
            continue

        target_cell_data = sheet.cell(row=row_match, column=p.WORKOUT_COLUMN).value

        if not target_cell_data:
            # success. Match found
            workouts_to_write.append(DataToWrite(title_datetime=workout_datetime,
                                                 data=workout.data,
                                                 target_row=row_match))

        elif target_cell_data == workout.data:
            workout_already_written.append(workout)

        elif target_cell_data != workout.data:
            # save the workout object, and existing cell contents, for later comparison / context
            target_cell_contains_clashing_info.append((workout, target_cell_data))

    # todo: review and simplify logic below
    # Processing done. Now alert user to different scenarios, and request user action if required
    if len(failed_to_find_date_cell) != 0:
        raise RuntimeError(f"Failed to find row matches for the following {len(failed_to_find_date_cell)} "
                           f"workouts. Please verify that the matching date value exists in the target Excel "
                           f"file, in the correct place.\n{failed_to_find_date_cell}")

    print(f"{len(workouts_to_write)} workouts can be written to target cells. {len(workout_already_written)} workouts "
          f"are already written to target cells")

    if len(target_cell_contains_clashing_info) != 0:
        # todo: make msg clearer
        print(f"The following {len(target_cell_contains_clashing_info)} workouts already have *different* values "
              f"written to their target cells. Please review")

        for workout, target_cell_data in target_cell_contains_clashing_info:
            neat_datetime = workout.title_datetime.strftime('%Y-%m-%d')
            similarity = uf.get_string_pct_similarity(workout.data, target_cell_data)
            print(f"{neat_datetime} INTENDED WRITE {similarity=}%:\t{workout.data}")
            print(f"{neat_datetime} EXISTING VALUE {similarity=}%:\t{target_cell_data}")

        inp = input("Do you wish to proceed, and OVERWRITE the existing values? (y/N) ")
        if inp.lower().strip() != "y":
            print("\nUser chose not to continue")
            exit()

    if len(workout_already_written) == len(parsed_workouts):
        print("No new workouts to write. Program exiting")
        exit()

    assert len(set([pd.target_row for pd in workouts_to_write])) == len(workouts_to_write), \
        "Error: multiple workouts are scheduled to be written to the same cell"
    return workouts_to_write


def write_data_to_xlsx(data_to_write: List[DataToWrite], backup=True) -> None:
    """
    Write data to the Excel file, optionally backing it up first. Use the WORKOUT_COLUMN and TARGET_PATH values
    specified in params.py. Perform minimal validation. Validation should be done prior to calling this function!
    :param data_to_write: a list of DataToWrite objects, containing the target row,
    :param backup: whether to back up the file before writing
    """

    assert all([isinstance(obj, DataToWrite) for obj in data_to_write])
    assert all([obj.target_row > 0 for obj in data_to_write]), "Invalid row for write object in function " \
                                                               "write_data_to_xlsx(...)"

    if backup:
        uf.backup_file_to_dir(file_name=p.TARGET_PATH, backup_directory=p.LOCAL_BACKUP_DIR)

    wb = openpyxl.load_workbook(p.TARGET_PATH)
    sheet = wb[p.TARGET_SHEET]

    print(f"Writing {len(data_to_write)} workouts to target file.")
    for packet in data_to_write:
        target_cell = sheet.cell(row=packet.target_row, column=p.WORKOUT_COLUMN)
        assert not target_cell.value, "Programming error. Target cell already has value written. No changes made"
        target_cell.value = packet.data

    wb.save(p.TARGET_PATH)


def is_date(string, fuzzy: bool = False) -> bool:
    """
    Return whether the string is likely to represent a date.
    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    # function inspired by Stackoverflow post
    try:
        # this is too liberal. For: 'July 23, day 3', it returns datetime.datetime(2003, 7, 23, 0, 0)
        parse(string, fuzzy=fuzzy)

    except (ValueError, OverflowError):
        # phone numbers can result in overflow in parse() function
        return False

    # reject strings like "17"
    if len(string) < 4:
        return False

    # for our purposes, a string must contain digits. Therefore, we reject strings like "September" as datelines,
    # but accept strings containing digits, such as "September 15" or "2 January"
    for c in string:
        if c.isdigit():
            return True
    return False


def capitalize_selectively(line: str) -> str:
    """
    Capitalize the first letter on each line. If the first letter in the string is an "x" following a digit, and
    followed by 2-3 digits, then don't capitalize anything.
    :param line: the string to process
    :return: the selectively capitalized string
    """
    reg = r'\dx\d\d?'
    if re.search(reg, line):
        return line

    else:
        for ind, c in enumerate(line):
            if c.isalpha():
                # capitalize the first letter
                return line[:ind] + line[ind].upper() + line[ind + 1:]
    return line


def line_is_comment(line: str) -> bool:
    """
    Return True if the input string matches the format used by comment lines, i.e. it starts with either "/" or "(".
    :param line: the string to process
    :return: True or False
    """
    if line.startswith('/'):
        return True
    if line.startswith('('):
        return True
    return False
