# retrieves bodyweights from Google Keep, then writes them to the correct row in the file specified by
# utilities.params.TARGET_PATH. Does some helpful things too, like alert the user to missing entries, etc
import gkeepapi.node
import openpyxl
import time
from datetime import datetime
from collections import UserDict
from typing import List, Union, Tuple, Any
import GKeepToCalc.utilities.params as p
import GKeepToCalc.utilities.utility_functions as uf


class RowBodyweightPairings(UserDict):
	# todo: add tests for this
	def __setitem__(self, row, bodyweight):
		# disallow duplicate rows, or the updating of values
		assert str(row) not in self.data.keys(), "This key has already been used"
		assert "." not in row, "No decimal places allowed in row"

		# validate bodyweight
		self.validate_bodyweight_text(bodyweight)
		self.data[int(row)] = str(bodyweight)

	@staticmethod
	def validate_bodyweight_text(bodyweight: str):
		legal_chars = list("0123456789?.")
		assert len(bodyweight) > 0, "Empty bodyweight received"
		assert bodyweight[0] != ".", f"Received invalid bodyweight {bodyweight}"

		illegal_remainder = bodyweight
		for legal_char in legal_chars:
			illegal_remainder = illegal_remainder.replace(legal_char, "")
		assert not illegal_remainder, f"Illegal characters found in bodyweight string {bodyweight}. Remainder=" \
									  f"'{illegal_remainder}'"


def return_most_recent_bodyweights(bodyweights: List[str], desired_count: int) -> List[str]:
	"""
	Return the X elements with the greatest index, stripped of spaces
	:param bodyweights: a list of strings
	:param desired_count: how many items to include in the returned list
	:return: a list of string bodyweights
	"""
	assert isinstance(desired_count, int), "Please provide an integer history_length"

	bodyweights = [weight for weight in bodyweights if weight]
	if len(bodyweights) == 0:
		return ['']

	if not all([isinstance(elem, str) for elem in bodyweights]):
		# we need string values because only strings can represent question marks "?" (i.e. absent bodyweights)
		raise ValueError(f"Error: non-string elements found in bodyweights list")

	# if history length < 0, then set it to 0
	desired_count = max(desired_count, 0)

	try:
		history = bodyweights[-desired_count:]
	except IndexError:
		# history greater than the number of bodyweights in Note
		history = bodyweights

	# strip spaces from each entry
	for ind, _ in enumerate(history):
		history[ind] = history[ind].replace(' ', '')

	return history


def format_bodyweight_history(history: List[str]) -> str:
	"""
	Take a string, parenthesize it, and comma-separate its elements. Include a trailing comma and space. If there's no
	history to format, then return the empty string
	:param history: a list of strings
	:return: a formatted string
	"""
	if len(history) == 0 or (len(history) == 1 and history[0] == ""):
		return ''
	return f"(" + ", ".join(history) + "), "


def trash_note_and_replace(keep, note, new_text) -> None:
	"""
	Trash input Note, and create new Note. (Items in trash remain available for 7 days, whereas changes to existing
	Notes may be irreversible. That's why we do not edit in place).
	We do not soft-code the title yet, as that's currently how we recognize the bodyweights note.
	:param keep: the Google Keep object
	:param note: the Note to be replaced
	:param new_text: the desired text of the new Note
	"""

	keep.createNote(title=p.BODYWEIGHTS_NOTE_TITLE, text=new_text)
	note.trash()
	keep.sync()
	print("Synchronizing")
	# without a wait sometimes sync doesn't complete
	time.sleep(3)


def find_bodyweights_note(notes: List[gkeepapi.node.Note]) -> gkeepapi.node.Note:
	"""
	Given a list of Notes, find the bodyweights note and return it. If multiple matching Notes are found, then raise a
	ValueError.
	:param notes: a list of Note objects through which to search
	:return: a Note object
	"""
	matches = []
	for note in notes:
		if note.trashed:
			continue

		if note.title.strip().lower() == p.BODYWEIGHTS_NOTE_TITLE.lower():
			matches.append(note)

	if len(matches) == 0:
		raise ValueError("No matching note found.\n"
						 "1) Does your bodyweight note exist?\n"
						 "2) Does it contain \"{p.BODYWEIGHTS_NOTE_TITLE}\" (without quotes) in its title?")

	if len(matches) > 1:
		raise ValueError(f"Several Notes found with \"{p.BODYWEIGHTS_NOTE_TITLE}\" in their title. Unable to determine"
						 f" which is the correct Note. Please trash the incorrect Note, or update the value of"
						 f" the bodyweights note title in params.py")
	return matches[0]


def return_note_edit_timestamp(bw_note: gkeepapi.node.Note) -> datetime.date:
	"""
	Return the edit time of the passed in Note object
	:param bw_note: the Note object
	:return: datetime object in form %Y-%m-%dT%H:%M:%S.%fZ (example: "2020-07-06 11:20:44.428000")
	"""
	return bw_note.timestamps.edited


def return_depunctuated_bodyweights_text(text,
										 keep_decimal_places=False,
										 keep_spaces=False,
										 keep_question_marks=False) -> str:
	"""
	Given an input string, return a copy with punctuation partially removed.
	:param text: the input string
	:param keep_decimal_places: if False, then remove decimal places from string
	:param keep_spaces: if False, then remove spaces from string
	:param keep_question_marks: if False, then remove question marks from string
	:return: a string with less punctuation
	"""
	txt = text.replace(",", "").replace("(", "").replace(")", "")
	if not keep_decimal_places:
		txt = txt.replace(".", "")
	if not keep_spaces:
		txt = txt.replace(" ", "")
	if not keep_question_marks:
		txt = txt.replace("?", "")
	return txt


def extract_bodyweights_from_validated_string(validated_str, split_on_parenthesis: bool) \
		-> Union[List, Tuple[List, List]]:
	"""
	Given a validated string, return the list of bodyweights found in that string. If split_on_parentheses,
	then return two lists - one containing values found inside parentheses, and one containing those outside
	:param validated_str: a validated string containing comma-separated floats, ints, "?", or parentheses etc
	:param split_on_parenthesis: whether to split on parentheses
	:return: one or two lists, containing bodyweights found in each group
	"""

	depunc_str = return_depunctuated_bodyweights_text(validated_str,
													  keep_decimal_places=True,
													  keep_spaces=True,
													  keep_question_marks=True)
	if len(depunc_str) == 0:
		# no bodyweights to return
		if split_on_parenthesis:
			return [], []
		return []

	if not split_on_parenthesis:
		# return all bodyweights in string
		lst = [val for val in depunc_str.split() if val.replace(" ", "")]
		return lst

	if split_on_parenthesis:
		# return 2 lists, split on closing parenthesis
		if ")" in validated_str:
			# this validation should already be done, but it's worth being certain
			assert validated_str.count(")") == 1, "Too many ')' found in string"
			validated_str = validated_str.replace("(", "").replace(",", "")
			split_1, split_2 = validated_str.split(")")
			context_window_weights = split_1.split()
			uncommitted_weights = split_2.split()
			return context_window_weights, uncommitted_weights
		else:
			bodyweights_lsts = depunc_str.split(), []
			return bodyweights_lsts


def validate_bodyweight_note_text(bw_note_text: str) -> None:
	"""
	If the bodyweight note text not as expected, then raise ValueError.
	:param bw_note_text: the string text found within the bodyweight note
	"""

	text = bw_note_text
	if "()" in text:
		# raise, to notify the user, in case (s)he accidentally ended up with "()" in the text
		raise ValueError("ERROR: empty parentheses in bodyweights note")

	if text.count("(") != text.count(")"):
		raise ValueError("ERROR: mismatched parentheses in bodyweights note")
	if text.count("(") > 1 or text.count(")") > 1:
		raise ValueError("ERROR: too many parentheses found in bodyweights note. Expected no more than 1 opening or "
						 "closing parenthesis respectively")
	# if text.count("(") == 1 and ")," not in text:
	# raise ValueError("ERROR: the context window is not followed by a comma")

	de_punctuated_text = return_depunctuated_bodyweights_text(text)
	if len(de_punctuated_text) > 0 and not de_punctuated_text.isdigit():
		raise ValueError("ERROR: bodyweights are incorrectly formatted. They should consist only of digits, with "
						 "1-2 optional decimal places, and each bodyweight should be followed by a comma")


def pair_new_bodyweights_with_rows(sheet, bodyweights_lst: [List, float], start_row: int, max_empty_rows=10) \
		-> RowBodyweightPairings:
	"""
	:param sheet: sheet in xlsx file containing bodyweights and dates
	:param bodyweights_lst: list of floats, representing bodyweights not yet committed to file
	:param start_row: the row at which the search starts
	:param max_empty_rows: the number of empty rows after which an error is raised, if a suitable row cannot be found.
	:return: a list of tuples, where tuple[0] is the int row to write to, and tuple[1] the str bodyweight. Empty cells
	are accounted for.
	"""

	assert isinstance(start_row, int)
	assert isinstance(max_empty_rows, int)

	current_row = start_row
	count_empty = 1
	pairings = RowBodyweightPairings()

	for bw in bodyweights_lst:
		date_cell_value = sheet.cell(row=current_row, column=p.DATE_COLUMN).value
		while date_cell_value is None:
			# skip empty cells in date column (e.g. at end of year), up to max length "max_empty_rows"
			current_row += 1
			count_empty += 1
			date_cell_value = sheet.cell(row=current_row, column=p.DATE_COLUMN).value
			if count_empty >= max_empty_rows:
				raise RuntimeError(f"Found too many empty date cells (the cutoff is {max_empty_rows}). "
								   f"Please verify that your date cell column contains enough non-empty values")

		# check if bodyweight cell is already written to
		bw_cell_value = sheet.cell(row=current_row, column=p.BODYWEIGHT_COLUMN).value
		if bw_cell_value is None:
			pairings[str(current_row)] = bw
			current_row += 1
		else:
			raise ValueError(f"Bodyweight cannot be written to target row {current_row} - cell already written to!"
							 f"No changes have been made")

	return pairings


def write_to_file(wb, sheet, row_bodyweight_pairings: RowBodyweightPairings) -> None:
	"""
	Write bodyweights to file. Expect validation to be done prior.
	:param wb: the workbook to write to
	:param sheet: the target sheet within the workbook
	:param row_bodyweight_pairings: the row and bodyweight pairings. A UserDict with validated entries
	"""

	for row, bodyweight in row_bodyweight_pairings.items():
		try:
			# we write as float because otherwise Calc (and perhaps Excel) prepend each value with a "'", to mark it as
			# a string, causing it to be left-aligned. The float conversion avoids that
			sheet.cell(row=row, column=p.BODYWEIGHT_COLUMN).value = float(bodyweight)
		except ValueError:
			# given bodyweight is probably "?"
			sheet.cell(row=row, column=p.BODYWEIGHT_COLUMN).value = bodyweight

	wb.save(p.TARGET_PATH)


def main():
	if not uf.target_path_is_xslx(p.TARGET_PATH):
		raise ValueError(f"Target path specified in params.py does not point to xlsx file. "
						 f"This is the path\n{p.TARGET_PATH}")
	if not uf.target_sheet_exists(p.TARGET_PATH, p.TARGET_SHEET):
		raise ValueError(f"Target xlsx does not contain sheet specified in params.py. "
						 f"This is the path\n{p.TARGET_PATH}")

	wb = openpyxl.load_workbook(p.TARGET_PATH)
	sheet = wb[p.TARGET_SHEET]
	keep = uf.login_and_return_keep_obj()
	notes = uf.retrieve_notes(keep)
	bw_note = find_bodyweights_note(notes)

	# if the user hasn't edited their bodyweights file recently, we do not write.
	bw_edit_timestamp = return_note_edit_timestamp(bw_note)

	# don't expect a bodyweight if run between 00:00 and 05:00
	# we also use this value to set our endpoint (final cell).
	today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
	if datetime.now().hour < 5:
		from datetime import timedelta
		today -= timedelta(days=1)

	if bw_edit_timestamp < today:
		print("- You have not edited your bodyweights note today.")
		print("- Please add today's bodyweight to the note. Then run the program again")
		print("- If you don't remember it, a question mark will be fine")
		print(f"bw_edit_timestamp={bw_edit_timestamp}, note text=\"{bw_note.text}\"")
		exit()

	start_row = uf.return_first_empty_bodyweight_row(sheet,
													 date_column=p.DATE_COLUMN,
													 bodyweight_column=p.BODYWEIGHT_COLUMN)
	todays_row = uf.find_row_of_cell_matching_datetime(sheet, today, date_column=p.DATE_COLUMN)

	if start_row == -1:
		raise ValueError("Start row not found")
	if todays_row == -1:
		raise ValueError("Failed to find the date cell corresponding to today's date in the xlsx file")
	if sheet.cell(todays_row, p.BODYWEIGHT_COLUMN).value:
		print("Value already written for today. Exiting program")
		exit()

	# check that text is as expected
	validate_bodyweight_note_text(bw_note.text)

	# alias the now-validated text
	validated_text = bw_note.text

	# separate bodyweights that have been committed to file form those that have not
	pre_existing_context_window, uncommitted_bodyweights = extract_bodyweights_from_validated_string(
		validated_text,
		split_on_parenthesis=True
	)

	if not uncommitted_bodyweights or (len(uncommitted_bodyweights) == 1 and uncommitted_bodyweights[0] == ""):
		print(f"Debug: Note.title='{bw_note.title}'; Note.text='{bw_note.text}'")
		print("INFO: no bodyweights found in Keep note. There is nothing new to write\nExiting")
		exit()

	# this is the number of bodyweights missing in the sheet, accounting for the fact that there may be empty rows
	# separating between target write cells (e.g. at year's end)
	# todo: simplify logic.
	num_expected_bodyweights = (todays_row - start_row + 1)
	count_empty_contiguous_rows = uf.count_empty_contiguous_rows_within_range(sheet=sheet,
																			  start_row=start_row,
																			  end_row=todays_row,
																			  cols_lst=[p.BODYWEIGHT_COLUMN])
	if num_expected_bodyweights != len(uncommitted_bodyweights):
		error_msg = f"Incorrect number of bodyweights supplied. Expected {num_expected_bodyweights} bodyweights in " \
					f"note. Found {len(uncommitted_bodyweights)} bodyweights. \nPlease correct the bodyweights note. " \
					f"If you've forgotten a value, then a question mark is a valid substitute for that bodyweight."
		raise ValueError(error_msg)

	if num_expected_bodyweights != count_empty_contiguous_rows:
		error_msg = f"{num_expected_bodyweights} bodyweights were provided in Google Keep Note. However, this does " \
					f"not match the number of empty rows found in the sheet. Please review the Excel file for stray " \
					f"values in the bodyweights column."
		raise ValueError(error_msg)

	# pair bodyweights with their target rows. Account for empty rows, and raise if anything is amiss.
	row_bodyweight_mapping: RowBodyweightPairings = pair_new_bodyweights_with_rows(sheet,
																				   uncommitted_bodyweights,
																				   start_row)

	# prepare history of the most recently committed-to-file bodyweights, for the bodyweight Note in Keep. This
	# may also be called the "context window".
	all_bodyweights = extract_bodyweights_from_validated_string(validated_text, split_on_parenthesis=False)
	most_recent_bodyweights: List[str] = return_most_recent_bodyweights(bodyweights=all_bodyweights,
																		desired_count=p.HISTORY_LENGTH)
	history: str = format_bodyweight_history(most_recent_bodyweights)

	uf.backup_target_path()
	print("Writing bodyweights to file")
	write_to_file(wb, sheet, row_bodyweight_mapping)

	# trash the bodyweights note, and replace it. The replacement has "history_length"
	# values saved in its history (i.e. context window)
	print("Updating note in Keep")
	trash_note_and_replace(keep, bw_note, history)
	print("Finished!")


if __name__ == '__main__':
	main()
