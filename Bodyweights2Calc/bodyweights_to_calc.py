# retrieves bodyweights from Google Keep, then writes them to the correct date cell in the specified file
# does intelligent stuff too, like alert the user to missing entries, etc
# consider creating a version for use by crontab
# REMEMBER to change params whenever necessary.

import openpyxl
import re
from datetime import datetime
import utilities.params as p
import utilities.utility_functions as uf


# todo: make program recognize duplicates?
# i.e. find the longest continuous string present in both Keep and the xlsx file
# why?
# in case there's overlap
# why does that matter?
#


def main():
    if not uf.target_is_xslx():
        raise ValueError("Target path specified in params.py does not point to xlsx file")
    if not uf.targetsheet_exists():
        raise ValueError("Target xlsx does not contain sheet specified in params.py")

    keep = uf.login_and_return_keep_obj()
    notes = uf.retrieve_notes(keep)
    bodyweights_lst = find_and_return_bodyweights(notes)

    wb = openpyxl.load_workbook(p.target_path)
    sheet = wb[p.target_sheet]

    # return range of rows requiring writes
    row_range_tpl = return_bw_rows_requiring_write(sheet)
    print(f'DEBUG: row_range_tpl={row_range_tpl}')
    print(f'DEBUG: bodyweights_lst={bodyweights_lst}')

    # confirm that the length of that range matches the number of bodyweights found in the Keep note
    if do_bodyweights_fill_all_vacancies(bodyweights_lst, row_range_tpl):
        uf.backup_targetpath()
        # writes to, but does not save file
        print("Writing to file")
        write_to_file(sheet, bodyweights_lst, row_range_tpl[0])
    else:
        # error messages already handled in condition function above
        exit()

    wb.save(p.target_path)

    # todo: write to file, then remove entries from Keep
    #   it's not a good idea to remove entries until we're certain that they're written
    #   so either don't remove them, or find a foolsafe way to make sure they're written
    print("DEV: entries will not be deleted from note after completion, because program is untested. It may fail")
    print("DEV: therefore you MUST remove verify that those bodyweights were written, then remove from Keep yourself")


def find_and_return_bodyweights(notes):
    """
    Within "notes", find the bodyweights note and return its modified contents
    :param notes: the Keep notes object (which contains all notes)
    :return: a list of integers
    """

    # we expect formats like these 3 below:
    # 83.2, 83, 83.4,
    # 101,
    # 100.4, 100.9, 99.8,
    # i.e. 2-3 digits with optional decimal place followed by a comma
    # spaces are optional. Commas are not. Each number must be followed by one comma
    bw_reg = re.compile(r'(\d{2,3}\.\d\s?,)+'
                        r'|(\d{2,3}\s?,)+'
                        r'|(\?{1,3}\s?,)+')

    for gnote in notes:
        # match either title or body. It's user preference where the weights will be
        for x in [gnote.title, gnote.text]:
            if not x.replace(",", "").replace(" ", "").replace(".", "").replace("?", "").isdigit():
                continue
            else:
                bodyweights = ["".join(m) for m in re.findall(bw_reg, x)]
                bodyweights = [t[:-1] for t in bodyweights]
                # This changes findall's output from this kind:
                # [('', '81,'), ('', '85,'), ('', '102,'), ('102.1,', '')]
                # to this kind
                # ['81', '85', '102', '102.1']

                if len(bodyweights) > 1:
                    return bodyweights

    raise ValueError("No matching note found. "
                     "1) Does your bodyweight note exist? "
                     "2) Is it in a valid format, with more than 1 entry? "
                     "3) Does it contain only numbers, spaces, commas and full stops?")


def return_bw_rows_requiring_write(sheet):
    """
    Return which rows should contain bodyweights but don't
    :param sheet: sheet in xlsx file containing bodyweights and dates
    :return: tuple of length 2, containing start and end rows
    """

    # keep going down bodyweight column,
    # counting empty bodyweight cells that neighbour a date cell
    # until date == today.
    # **we assume that every date is intended to have an accompanying bodyweight**
    start = None
    count_unwritten_cells = 0

    for t in range(1, 1000000):
        # some rows in this column are strings
        if isinstance(sheet.cell(row=t, column=p.date_column).value, datetime):
            if sheet.cell(row=t, column=p.date_column).value > datetime.now():
                return start, start + count_unwritten_cells
        if sheet.cell(row=t, column=p.bodyweight_column).value is None:
            if isinstance(sheet.cell(row=t, column=p.date_column).value, datetime):
                # empty bodyweight cell found next to a date cell.
                if not start:
                    start = t
                else:
                    count_unwritten_cells += 1


def do_bodyweights_fill_all_vacancies(bodyweights_lst, row_range):
    """
    :param bodyweights_lst: list of bodyweights in such a format: ['81', '85', '102', '102.1']
    :param row_range: tuple containing number of first and last rows lacking bodyweights
    :return: True if the number of provided bodyweights equals the number of absent bodyweights. Else False.
    """
    # informs user if values are missing or too numerous
    # relative to the vacancies present in the xlsx file

    # number of empty cells. 1+ makes it inclusive
    count_required = 1 + row_range[1] - row_range[0]
    # number of provided bodyweights. Skip first value
    count_provided = len(bodyweights_lst) - 1

    if count_provided < count_required:
        print(f"Too few values provided. Needed {count_required}, provided with {count_provided}")
        return False
    elif count_provided > count_required:
        print(f"Too many values provided. Needed {count_required}, provided with {count_provided}")
        return False
    return True


def write_to_file(sheet, bodyweights_lst, start_row):
    """
    :param sheet: sheet in xlsx file containing bodyweights and dates
    :param bodyweights_lst: list of bodyweights in such a format: ['81', '85', '102', '102.1']
    :param start_row (int)
    """
    for bw in bodyweights_lst:
        if sheet.cell(row=start_row, column=p.bodyweight_column).value is None:
            # TODO: resolve
            # for mysterious reasons, every weight gets written with a prepended '
            # is this an openpyxl bug? .replace("'","") does nothing to fix problem
            '''
            says: https://superuser.com/questions/394092/how-to-remove-a-plain-text-protecting-single-quote-from-all-the-selected-cells-i
            
            You can remove the leading single quote 
            (which actually isn't part of the string in the cell) 
            using a regex-based search and replace:

            Search for all characters between the start and end of the string 
            ^.*$
            replace with match 
            &
            
            For some reason, in LibreOffice, "Data" menu -> "Text to columns" also works
            ...at least after the fact. Idk whether it's preventative.
            
            the apostrophe is an indicator that a cell is formatted as numeric/date value
            and apparently does not change the value of the cell.
            but it does left adjust it, which is irritating.
            
            Did you try sheet.cell("C1").set_explicit_value("value", 's')
            '''
            sheet.cell(row=start_row, column=p.bodyweight_column).value = bw
            start_row += 1
        else:
            print(f"Cannot write to cell {start_row} - cell already written to!")
            print("No changes have been made")
            exit()


if __name__ == '__main__':
    main()
