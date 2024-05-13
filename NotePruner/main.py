from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import Dict, List
from collections import Counter

import openpyxl
from tabulate import tabulate

import utilities.params as p
import utilities.utility_functions as uf
from utilities.shared_types import Entry


@dataclass
class DiscardCandidate:
    floored_date: datetime
    note: Entry
    in_sheet_as: str

    def __post_init__(self):
        assert isinstance(self.floored_date, datetime)
        assert isinstance(self.note, Entry)
        assert isinstance(self.in_sheet_as, str)
        self.floored_date = self.floored_date.replace(hour=0, minute=0, second=0, microsecond=0)


def retrieve_note_snippets_from_xlsx(sheet,
                                     workout_notes: List[Entry]) -> Dict[datetime, str]:
    # Retrieve workout column values (including the empty string) from the target xlsx file, up to a given row.
    # Return them as a dictionary, where each key is a datetime object representing the date of the workout (without
    # time component), and each value is the workout's string

    start_date = min([note.floored_datetime for note in workout_notes])
    end_date = max([note.floored_datetime for note in workout_notes])

    start_row = uf.find_row_of_cell_matching_datetime(sheet, datetime_target=start_date, date_column=p.DATE_COLUMN)
    end_row = uf.find_row_of_cell_matching_datetime(sheet, datetime_target=end_date, date_column=p.DATE_COLUMN)

    xlsx_snippets = dict()
    for row in sheet.iter_rows(min_row=start_row,
                               min_col=0,  # we don't exclude columns because that throws our indexing off
                               max_col=max(p.WORKOUT_COLUMN, p.DATE_COLUMN),
                               max_row=end_row,
                               values_only=False):
        # a "row" in this context is a tuple of cells. We subtract 1 because openpyxl columns are 1-indexed, but
        # iter_rows is returning a tuple, which is 0-indexed.
        date_value = row[p.DATE_COLUMN - 1].value

        if not date_value:
            # if there's no date, there's no snippet to store
            continue

        date_value = uf.convert_string_to_datetime(date_value, regress_future_dates=False)
        floored_date = date_value.replace(hour=0, minute=0, second=0)
        assert not xlsx_snippets.get(floored_date), (f"Multiple workouts found for date {floored_date} in the target "
                                                     f"file. This is not a supported use case.")
        xlsx_snippets[floored_date] = row[p.WORKOUT_COLUMN - 1].value

    return xlsx_snippets


def get_discard_candidates(sheet,
                           workout_notes: List[Entry],
                           end_date: datetime) -> List[DiscardCandidate]:
    xlsx_snippets: Dict[datetime, str] = retrieve_note_snippets_from_xlsx(sheet=sheet,
                                                                          workout_notes=workout_notes)

    discard_candidates = []
    for note in workout_notes:
        floored_date = note.floored_datetime
        if (in_sheet_as := xlsx_snippets.get(floored_date)) and is_discard_candidate(note=note,
                                                                                     xlsx_snippets=xlsx_snippets,
                                                                                     end_date=end_date):
            discard_candidates.append(DiscardCandidate(floored_date=floored_date, note=note, in_sheet_as=in_sheet_as))

    return discard_candidates


def is_discard_candidate(note: Entry, xlsx_snippets: Dict[datetime, str], end_date: datetime) -> bool:
    # todo: test this properly
    """
    Return True if a note qualifies as being ready for discarding, else False.
    For a given note to return True, the following criteria must be met:
     1) the note is already written to the target file...
     2) ...for the correct date
     3) that date is earlier than the requested end date.

    :param xlsx_snippets: a dictionary where each key is the date a workout took place, and each value its string
    representation in the target Excel file. This date should be without a time component.
    :param note: the note we're considering as a discard candidate.
    :param end_date: the cutoff point, after which we always return False
    """

    # It only makes sense to discard workout notes
    if not note.is_valid_workout_note():
        return False

    # do not flag as discard candidate beyond cutoff
    assert isinstance(end_date, datetime)
    if note.floored_datetime and note.floored_datetime > end_date.replace(hour=0, minute=0, second=0, microsecond=0):
        return False

    # is note already written to the target file?
    return bool(xlsx_snippets.get(note.floored_datetime))


def present_discard_candidates(discard_candidates: List[DiscardCandidate]) -> None:
    """
    Present discard candidates to the user, in table format, demonstrating: the discard candidates, their
    corresponding values in the passed-in dictionary, and a percentage similarity rating of the two strings.
    :param discard_candidates: a list of notes considered eligible for discarding.
    representation in the target xlsx file.

    TABLE FORMATTED AS BELOW:

    date    note to be discarded snippet      snippet from xlsx
    ...     ...                             ...
    ...     ...                             ...

    Discard? (Y/N)
    """
    print(f"These are the {len(discard_candidates)} discard candidates. They are already written to file, and "
          "are older than your specified date range")
    print("\n**DISCARD CANDIDATES**")

    # populate the matrix for tabulate
    tabulate_matrix = []
    sorted_discard_candidates = sorted(discard_candidates, key=lambda x: x.floored_date)
    for candidate in sorted_discard_candidates:
        # remove comment lines and Obsidian properties, because they don't appear in the xlsx file,
        # and are therefore unhelpful for a side-by-side comparison
        text_no_properties = uf.strip_obsidian_properties(candidate.note.text)
        note_snippet = return_note_text_minus_comments(text_no_properties, remove_plus_signs=True).replace('\n', ' ')
        note_snippet = note_snippet[:p.SNIPPET_LENGTH]

        floored_date = candidate.floored_date
        printable_date = floored_date.strftime("%Y-%m-%d")
        xlsx_snippet = candidate.in_sheet_as.rstrip()[:p.SNIPPET_LENGTH]
        similarity = uf.get_string_pct_similarity(note_snippet, xlsx_snippet)

        # append the table row
        tabulate_matrix.append([printable_date, note_snippet, xlsx_snippet, str(similarity) + "%"])

    headers = ["Date", "Note snippet", "Exists in xlsx as...", "Similarity"]
    print(tabulate(tabulate_matrix, headers=headers))
    print()


def is_discard_requested() -> bool:
    # returns True if permission is given to discard ALL notes presented by present_discard_candidates()
    discard_requested = input("Discard all of the above? (y/N): ").strip().lower()
    return discard_requested in ["y", "yes"]


def greet() -> None:
    greeting = "\n\tNOTE PRUNER \n" + \
               "\tdiscards workout notes from the source location, up to an optionally specified user-given date\n" \
               "\tprovided they are already written to file and the user approves\n"
    print(greeting)


def request_end_date() -> datetime:
    print("\tTo start, enter the inclusive date up to which you wish to discard notes (an inclusive count). Either "
          "press enter to use today's date, or enter another date in DDMM format (e.g. `2305`). Dates in the future "
          "will be regressed to last year.")
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    while True:

        end_date = input(">").replace(' ', '')
        if not end_date:
            return today

        while not end_date.isdigit():
            print("Please enter a valid date")
            end_date = input('>').replace(' ', '')

        # assume current year unless that date is in the future
        target_date = end_date + str(today.year)
        try:
            target_date = datetime.strptime(target_date, '%d%m%Y')
        except ValueError as e:
            print("Error:", e)
            print("Unable to parse given date")
            continue
        if target_date > datetime.now():
            # We account for leap years
            target_date -= timedelta(days=365 + 1 - today.year % 4)
        print(datetime.strftime(target_date, "%d/%m/%Y"))

        response = input(">Is this date correct? (y/N)").lower()
        if response.lower().strip() in ["y", "yes"]:
            return target_date


def return_note_text_minus_comments(note_text: str, remove_plus_signs=False) -> str:
    # given a note, return its text as a string, with comment lines omitted
    retstr = ''
    for line in note_text.split('\n'):
        line = line.lstrip().replace('\n', '')
        if line.startswith(('/', '(')):
            continue
        if remove_plus_signs and len(line) > 2:
            # remove "+" because it's not relevant for comparisons in present_discard_candidates(...)
            retstr += line.replace('+ ', '').replace('+', '') + ' '

    return retstr


def main():
    uf.validate_target_sheet_params()

    # fail early: try this before greeting the user, in case that it fails (e.g. because of user config problem)
    handler = uf.return_handler()
    notes = handler.retrieve_notes()
    workout_notes = [note for note in notes if note.is_valid_workout_note()]
    if not workout_notes:
        print("No workout notes found. Nothing to prune. Program exiting")
        exit()

    repeated_dates = [dt for dt, count in Counter([note.floored_datetime for note in workout_notes]).items()
                      if count > 1]
    if repeated_dates:
        raise ValueError(f"Multiple workout notes found for the same date. This is not a supported use case. "
                         f"Please ensure that each workout has a unique date. Offending dates: {repeated_dates=}")

    # source and target are ready. Request user input.
    greet()
    end_date = request_end_date()

    wb = openpyxl.load_workbook(p.TARGET_PATH)
    sheet = wb[p.TARGET_SHEET]

    discard_candidates: List[DiscardCandidate] = get_discard_candidates(sheet, workout_notes, end_date)
    present_discard_candidates(discard_candidates=discard_candidates)

    if not discard_candidates:
        print("No notes found to discard. Program exiting")
        exit()

    if not is_discard_requested():
        print("No changes made")
        exit()

    certain = input("Press 'C' to confirm discard. Any other key to undo: ").lower().strip()
    if certain != 'c':
        print("No changes made")
        exit()
    else:
        handler.discard_notes([candidate.note for candidate in discard_candidates])
        print("Specified notes discarded. Program execution complete.")


if __name__ == '__main__':
    main()
