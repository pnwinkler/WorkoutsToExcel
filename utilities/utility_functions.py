import os
import shutil
from datetime import datetime
from difflib import SequenceMatcher
from typing import List

import openpyxl

import utilities.params as p
from utilities.shared_types import Handler


def validate_target_sheet_params() -> None:
    if not target_path_is_xslx(p.TARGET_PATH):
        raise ValueError(f"Target path specified in params.py does not point to xlsx file. "
                         f"This is the path\n{p.TARGET_PATH}")
    if not target_sheet_exists(p.TARGET_PATH, p.TARGET_SHEET):
        raise ValueError(f"Target xlsx does not contain sheet specified in params.py. "
                         f"This is the path\n{p.TARGET_PATH}")


def return_handler() -> Handler:
    match p.RETRIEVAL_METHOD:
        case p.GKEEPAPI_STR:
            import utilities.keep_api_handler as kf
            return kf.KeepApiHandler()
        case p.LOCAL_STR:
            import utilities.local_file_handler as lr
            return lr.LocalFileHandler()
        case _:
            raise NotImplementedError(f"Retrieval method {p.RETRIEVAL_METHOD} not implemented.")


def backup_file_to_dir(source_file_path: str,
                       backup_directory: str,
                       basename_override: str = "",
                       keep_date_info=True) -> None:
    """
    This function backs up a file to a specified directory. If the directory does not exist, it creates it. The
    default format of the new basename is "backup_YYYY_MM_DD_source_file_name". However, this can be overridden by
    passing in the correct parameters.

    :param source_file_path: The full path to the file to be backed up.
    :param backup_directory: The directory to back the file up to. The path must be a full path.
    :param basename_override: An optional string to override the basename of the backup file. Unless keep_date_info is
    set to True, then this string will be the full basename.
    :param keep_date_info: A boolean indicating whether to include the current date in the backup file's name.

    :return: None.
    """
    os.makedirs(backup_directory, exist_ok=True)

    basename_parts = []
    if not basename_override:
        basename_parts.append('backup')
    else:
        basename_parts.append(basename_override)

    if keep_date_info:
        now = datetime.now()
        ymd = '_'.join(str(v) for v in [now.year, now.month, now.day])
        basename_parts.append(ymd)

    if not basename_override:
        basename_parts.append(os.path.basename(source_file_path))

    extension = os.path.splitext(source_file_path)[1]
    full_backup_path = os.path.join(backup_directory, "_".join(basename_parts)) + extension
    shutil.copy(source_file_path, full_backup_path)


def convert_string_to_datetime(date_str: str, regress_future_dates=True) -> datetime:
    """
    Return the input string's datetime equivalent. Raise on failure to convert.
    :param date_str: the string to convert
    :param regress_future_dates: if true, then subtract one year from the date to be returned, if that date is in the
    future as of the time of execution.
    :return: a datetime object
    """
    if isinstance(date_str, datetime):
        return date_str

    assert isinstance(date_str, str), f"Invalid parameter type received {type(date_str)}. Expected string"
    for char in ['\n', ';', ' ', '.', '-', '_', '/']:
        date_str = date_str.replace(char, '')

    # try to match the date string to a datetime object, with and without year
    for year_format in ['%Y%m%d', '%d%B%Y', '%d%b%Y', '%B%d%Y', '%b%d%Y', '%d%B', '%d%b', '%B%d', '%b%d']:
        try:
            datetime_obj = datetime.strptime(date_str, year_format)
        except ValueError:
            continue

        now = datetime.now()
        if datetime_obj.year < 2000:
            # year was not specified in the date string. Assume it's the current year.
            datetime_obj = datetime_obj.replace(year=now.year)

        if now < datetime_obj and regress_future_dates:
            # datetime is in the future, but future date is not wanted. Return previous year.
            return datetime_obj.replace(year=now.year - 1)
        return datetime_obj

    # matching to datetime failed, both with and without year
    raise ValueError(f"Failed to convert this value to datetime: '{date_str}'")


def count_empty_contiguous_rows_within_range(sheet, start_row: int, end_row: int, cols_lst: List[int]) -> int:
    """
    Return an inclusive count of the contiguously empty rows between start and end rows, where all cells in each of
    those rows are empty, for all columns in the columns list.
    :param sheet: the Excel sheet
    :param start_row: the row at which to start counting
    :param end_row: the final row to check
    :param cols_lst: the columns in which to check for values
    :return: a count of the empty rows found
    """
    # todo: consider removing this. It's only used in 1 place so far
    if isinstance(cols_lst, str):
        cols_lst = list(cols_lst)
    cols = [int(x) for x in cols_lst]

    count = 0
    for row in range(start_row, end_row + 1):
        for col in cols:
            if sheet.cell(row=row, column=col).value:
                return count
        count += 1
    return count


# todo: refactor
def find_row_of_cell_matching_datetime(sheet,
                                       datetime_target: datetime.date,
                                       date_column: int | str,
                                       raise_on_failure=False) -> int:
    """
    Returns row value of cell containing specified date in specified column. Returns -1 if not found
    :param sheet: an Excel sheet object
    :param datetime_target: the datetime date to search for in the date_column
    :param date_column: the column in which to search for date
    :param raise_on_failure: whether to raise a RuntimeError or return -1 on failure to find matching date
    :return: row number or -1 or RuntimeError
    """
    datetime_target = datetime_target.replace(hour=0, minute=0, second=0, microsecond=0)

    # convert date_column == 1 into "A", for example.
    col_letter = chr(date_column + 64)

    # find date cell matching the "date" parameter in the given sheet
    # note that in xlsx files: headers and strings are str, dates are datetime objects, empty cells are NoneType
    for cell in sheet[col_letter]:
        if isinstance(cell.value, datetime) and cell.value == datetime_target:
            return cell.row

    if raise_on_failure:
        err_msg = f"Failed to find matching date cell in target sheet, column {col_letter}"
        raise RuntimeError(err_msg)
    return -1


def return_first_absent_bodyweight_row(sheet, date_column: int, bodyweight_column: int) -> int:
    """
    Find the smallest row number, where:
     1) said row contains a date string in the date column
     2) said row contains no bodyweight in the bodyweights column
     3) the row above said row contains both date and bodyweight values.
    :param sheet: the Excel sheet
    :param date_column: the column in which date values are saved
    :param bodyweight_column: the column in which bodyweights are saved
    :return: an integer, representing a row number
    """

    todays_row = find_row_of_cell_matching_datetime(sheet, datetime.now(), date_column, raise_on_failure=True)
    if sheet.cell(row=todays_row, column=bodyweight_column).value:
        raise RuntimeError(f"Today's bodyweight cell is already written to")

    first_occurrence = None
    for row in range(todays_row, 0, -1):
        # search backwards, for performance reasons
        date_cell_value = sheet.cell(row=row, column=date_column).value
        bw_cell_value = sheet.cell(row=row, column=bodyweight_column).value
        row_has_date = isinstance(date_cell_value, datetime)
        row_has_bodyweight = isinstance(bw_cell_value, (str, float, int))

        if row_has_date and not row_has_bodyweight:
            # suitable candidate
            first_occurrence = row
        if row_has_date and row_has_bodyweight:
            # we've reached the previously filled in row.
            if first_occurrence:
                return first_occurrence

            # we found no matches
            break

    raise ValueError("Failed to find empty bodyweight cell.")


def target_path_is_xslx(file_path: str) -> bool:
    filename, extension = os.path.splitext(file_path)
    return extension in ['.xlsx', '.xls']


def target_sheet_exists(excel_path: str, target_sheet_name: str) -> bool:
    """
    Return True if target sheet is found in Excel at path, else False.
    :param excel_path: a string path pointing to an Excel file
    :param target_sheet_name: a sheet name
    :return: True / False
    """
    wb = openpyxl.load_workbook(excel_path)
    return target_sheet_name in wb.sheetnames


def get_string_pct_similarity(str_1, str_2) -> int:
    float_num = SequenceMatcher(None, str_1, str_2).ratio()
    return int(float_num * 100)


def strip_obsidian_properties(text: str) -> str:
    """
    Given a string, remove any Obsidian properties from it, then return it
    :param text: the string to examine
    :return: the string with Obsidian properties stripped out
    """

    # Obsidian uses these 3 dashes to indicate the start and end of a properties section
    obsidian_separator = "---"

    if not text.startswith(obsidian_separator):
        return text

    splt = text.split('\n')[1:]
    for idx, line in enumerate(splt):
        if line.strip() == obsidian_separator:
            return '\n'.join(splt[idx + 1:])

    raise ValueError("Found Obsidian separator only at start of note, but it's expected to occur twice")
