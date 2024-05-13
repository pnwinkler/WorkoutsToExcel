from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Tuple

import openpyxl

import utilities.params as p
import utilities.utility_functions as uf
from utilities.shared_types import Entry


@dataclass
class ParsedWorkout:
    # title_datetime is the interpreted date time of the note's title
    title_datetime: datetime

    # the formatted workout data
    data: str

    def __post_init__(self):
        assert isinstance(self.title_datetime, datetime)
        assert isinstance(self.data, str)
        # we don't want to accept workouts without a clearly defined year. 2000 is an arbitrary cutoff
        assert self.title_datetime.year > 2000, "Workout year must be specified"

    def __repr__(self):
        return f"<{self.title_datetime}>: {self.data}"


def parse_workout_notes(workout_notes: List[Entry]) -> List[ParsedWorkout]:
    """
    Given a list of workout notes, clean up and format the workout within each note, then return it as a list
    of ParsedWorkout objects, each representing one workout in its final format.
    :param workout_notes: a list of notes, each representing a workout. Title dates should be unique.
    :return: a list of ParsedWorkout objects
    """

    for note in workout_notes:
        assert note.is_valid_workout_note()

    parsed_data_lst = []
    for note in workout_notes:
        # strip lines, remove Obsidian properties, drop empty lines and comment lines.
        raw_text_no_properties: str = uf.strip_obsidian_properties(note.text)
        workout_text = [line.strip() for line in raw_text_no_properties.split('\n')
                        if line
                        and not (line_is_comment(line) or line.startswith('\n'))]

        # clean up and format one workout
        single_workout_lines = []
        for line in workout_text:
            parsed_line = capitalize_selectively(line)

            # clean-up note text
            for char in [';', '..', ' .']:
                parsed_line = parsed_line.replace(char, '.')
            parsed_line = parsed_line.replace('\n', '')
            parsed_line = parsed_line.replace('  ', ' ')

            # the "+" symbol can be used at the beginning of the line in a note, to indicate an "extra" exercise (i.e.
            # one not part of the standard workout). We include the line, but not the "+" symbol.
            parsed_line = parsed_line.lstrip('+ ')
            parsed_line = parsed_line.lstrip('+')

            # trailing semi-colons can happen due to data entry errors
            if parsed_line.endswith(":"):
                parsed_line = parsed_line[:-1]

            parsed_line = parsed_line.rstrip()

            # add that cleaned line to the workout's lines
            single_workout_lines.append(parsed_line)

        # semicolon-separate each exercise
        exercises_str = '; '.join(single_workout_lines)

        # replace the final semicolon with a full stop.
        exercises_str, est_xx_mins_line = exercises_str.rsplit('; ', 1)
        complete_workout_text = exercises_str + ". " + est_xx_mins_line

        # save the formatted workout
        parsed_data_lst.append(ParsedWorkout(title_datetime=note.floored_datetime, data=complete_workout_text))

    return parsed_data_lst


def pair_workouts_with_rows(target_sheet, parsed_workouts: List[ParsedWorkout]) -> Dict[int, ParsedWorkout]:
    """
    Given a list of parsed workouts, pair each workout with a unique row in the target file, such that the cell value
    in the date column of that row equals the value of the workout's interpreted datetime.
    :param target_sheet: the target sheet to inspect
    :param parsed_workouts: a list of fully formatted workouts
    :return: a list of parsed workouts, each paired with suitable row number.
    """
    if not len(parsed_workouts):
        print("No workouts to write")
        exit()

    # the object to be returned
    workouts_to_write: Dict[int, ParsedWorkout] = {}

    # collect errors
    failed_to_find_date_cell: [ParsedWorkout] = []
    workout_already_written: [ParsedWorkout] = []
    workout_info_clashes: Dict[int, Tuple[ParsedWorkout, str]] = {}

    sheet = target_sheet
    for workout in parsed_workouts:
        row_match = uf.find_row_of_cell_matching_datetime(sheet=sheet,
                                                          datetime_target=workout.title_datetime,
                                                          date_column=p.DATE_COLUMN,
                                                          raise_on_failure=False)
        if row_match == -1:
            failed_to_find_date_cell.append(workout)
            continue

        target_cell_data = sheet.cell(row=row_match, column=p.WORKOUT_COLUMN).value
        if not target_cell_data:
            # success. Match found and cell is empty
            assert row_match not in workouts_to_write.keys(), ("Error: multiple workouts are scheduled to be written "
                                                               f"to the same cell, in row {row_match}")
            assert row_match > 0
            workouts_to_write[row_match] = workout

        elif target_cell_data == workout.data:
            workout_already_written.append(workout)

        else:
            # save the workout object, and existing cell contents, for later comparison / context
            workout_info_clashes[row_match] = (workout, target_cell_data)

    # processing done
    if len(failed_to_find_date_cell) != 0:
        raise RuntimeError(f"Failed to find row matches for the following {len(failed_to_find_date_cell)} "
                           f"workouts. Please verify that each of the matching date value exist in the target Excel "
                           f"file, in the correct place.\n{failed_to_find_date_cell}")

    print(f"{len(workouts_to_write)} new workouts can be written to target cells. "
          f"{len(workout_already_written)} workouts are already written to target cells")

    if len(workout_already_written) == len(parsed_workouts):
        print("No new workouts to write. Program exiting")
        exit()

    if len(workout_info_clashes) != 0:
        print(f"The following {len(workout_info_clashes)} workouts already have *different* values "
              f"written to their target cells in the Excel.")

        for workout, target_cell_data in workout_info_clashes.values():
            neat_datetime = workout.title_datetime.strftime('%Y-%m-%d')
            similarity = uf.get_string_pct_similarity(workout.data, target_cell_data)
            print(f"{neat_datetime} INTENDED WRITE {similarity=}%:\t{workout.data}")
            print(f"{neat_datetime} EXISTING VALUE {similarity=}%:\t{target_cell_data}")

        inp = input("Do you wish to proceed, and OVERWRITE the existing values? (y/N) ")
        if inp.lower().strip() not in ["y", "yes"]:
            print("\nUser chose not to continue")
            exit()

    conflicting_workouts = {row: v[0] for row, v in workout_info_clashes.items()}
    # sanity checks
    assert all(isinstance(workout, ParsedWorkout) for workout in conflicting_workouts.values())
    assert set(workouts_to_write.keys()).isdisjoint(set(conflicting_workouts.keys()))
    return workouts_to_write | conflicting_workouts


def write_data_to_xlsx(data_to_write: Dict[int, ParsedWorkout], backup=True) -> None:
    """
    Write data to the Excel file. Back it up first if requested. Validation should be done prior to calling this
    function.
    :param data_to_write: a dict of objects, where the key is the target row, and the value the string to write
    :param backup: whether to back up the file before writing
    """

    if backup:
        uf.backup_file_to_dir(source_file_path=p.TARGET_PATH, backup_directory=p.LOCAL_EXCEL_BACKUP_DIR)

    wb = openpyxl.load_workbook(p.TARGET_PATH)
    sheet = wb[p.TARGET_SHEET]

    print(f"Writing {len(data_to_write)} workouts to target file.")
    for row, workout in data_to_write.items():
        target_cell = sheet.cell(row=row, column=p.WORKOUT_COLUMN)
        target_cell.value = workout.data

    wb.save(p.TARGET_PATH)


def capitalize_selectively(line: str) -> str:
    """
    Capitalize the first letter on each line.
    :param line: the string to process
    :return: the processed string
    """
    # If the first letter in the string is an "x" following a digit, and followed by 2-3 digits (for example "3x10" or
    # "5x3", then don't capitalize anything.
    # reg = r'\dx\d\d?'
    # if re.search(reg, line):
    #     return line

    # else:
    for ind, c in enumerate(line):
        if c.isalpha():
            # capitalize the first letter
            return line[:ind] + line[ind].upper() + line[ind + 1:]
    return line


def line_is_comment(line: str) -> bool:
    """
    Return True if the input string matches the format used by comment lines, i.e. it starts with either "/" or "(".
    :param line: the string to process
    :return: True or False
    """
    if line.startswith('/'):
        return True
    if line.startswith('('):
        return True
    return False
