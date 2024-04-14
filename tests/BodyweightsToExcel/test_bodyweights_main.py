import unittest
from openpyxl import Workbook
from utilities.params import DATE_COLUMN, BODYWEIGHT_COLUMN
from BodyweightsToExcel.main import *


class TestPairNewBodyweightsWithRows(unittest.TestCase):
    def setUp(self):
        self.sheet = Workbook().active
        self.bodyweights = [70.5, 71.2, 70.8, 70.6]
        # This is the row where the first date cell is found. We deliberately choose an incorrect value (the row of
        # the column's title), to ensure that the function can handle this.
        self.start_row = 1
        self.max_empty_rows = 10

        self.sheet.cell(row=1, column=DATE_COLUMN).value = "Date column title"
        self.sheet.cell(row=2, column=DATE_COLUMN).value = "2021-01-01"
        self.sheet.cell(row=3, column=DATE_COLUMN).value = "2021-01-02"
        self.sheet.cell(row=4, column=DATE_COLUMN).value = "2021-01-03"
        self.sheet.cell(row=5, column=DATE_COLUMN).value = "2021-01-04"

    def test_pairs_bodyweights_with_correct_rows(self):
        result = pair_new_bodyweights_with_rows(self.sheet, self.bodyweights, self.start_row, self.max_empty_rows)
        self.assertIsInstance(result, RowBodyweightPairings)
        self.assertEqual(len(result), len(self.bodyweights))
        self.assertEqual(float(result[2]), 70.5)
        self.assertEqual(float(result[5]), 70.6)

    def test_raises_error_when_too_many_empty_date_cells(self):
        self.max_empty_rows = 0
        with self.assertRaises(RuntimeError):
            pair_new_bodyweights_with_rows(self.sheet, self.bodyweights, self.start_row, self.max_empty_rows)

    def test_raises_error_when_bodyweight_cell_already_written(self):
        self.sheet.cell(row=2, column=BODYWEIGHT_COLUMN).value = 70.0
        with self.assertRaises(RuntimeError):
            pair_new_bodyweights_with_rows(self.sheet, self.bodyweights, self.start_row, self.max_empty_rows)

    def test_handles_empty_bodyweights_list(self):
        self.bodyweights = []
        result = pair_new_bodyweights_with_rows(self.sheet, self.bodyweights, self.start_row, self.max_empty_rows)
        self.assertIsInstance(result, RowBodyweightPairings)
        self.assertEqual(len(result), 0)


class TestRowBodyweightPairings(unittest.TestCase):
    def setUp(self):
        self.row_bodyweight_pairings = RowBodyweightPairings()

    def test_add_valid_bodyweight(self):
        self.row_bodyweight_pairings[1] = 70.5
        self.assertEqual(self.row_bodyweight_pairings[1], '70.5')

    def test_add_invalid_bodyweight(self):
        with self.assertRaises(AssertionError):
            self.row_bodyweight_pairings[1] = '70a'

    def test_add_duplicate_row(self):
        self.row_bodyweight_pairings[1] = 70.5
        with self.assertRaises(AssertionError):
            self.row_bodyweight_pairings[1] = 71.5


class TestReturnMostRecentBodyweights(unittest.TestCase):
    def test_return_most_recent_bodyweights(self):
        bodyweights = ['70.5', '71.2', '70.8', '70.6']
        result = return_most_recent_bodyweights(bodyweights, 2)
        self.assertEqual(result, ['70.8', '70.6'])

    def test_return_most_recent_bodyweights_with_empty_list(self):
        bodyweights = []
        result = return_most_recent_bodyweights(bodyweights, 2)
        self.assertIsNone(result)


class TestFormatBodyweightHistory(unittest.TestCase):
    def test_format_bodyweight_history(self):
        history = ['70.5', '71.2']
        result = format_bodyweight_history(history)
        self.assertEqual(result, '(70.5, 71.2), ')

    def test_format_bodyweight_history_with_empty_list(self):
        history = []
        result = format_bodyweight_history(history)
        self.assertEqual(result, '')


class TestExtractBodyweightsFromString(unittest.TestCase):
    def test_extract_bodyweights_from_string(self):
        raw_string = '(70.5,71.2), ?, 70.6'
        result = extract_bodyweights_from_string(raw_string, True)
        self.assertEqual(result, (['70.5', '71.2'], ['?', '70.6']))

    def test_extract_bodyweights_from_string_with_no_parentheses(self):
        raw_string = '70.5, 71.2,?, 70.6'
        result = extract_bodyweights_from_string(raw_string, True)
        self.assertEqual(result, ([], ['70.5', '71.2', '?', '70.6']))


class TestValidateBodyweightNoteText(unittest.TestCase):
    def test_validate_bodyweight_note_text(self):
        bw_note_text = '(70.5, 71.2), 70.8, 70.6'
        try:
            validate_bodyweight_note_text(bw_note_text)
        except ValueError:
            self.fail("_validate_bodyweight_note_text raised ValueError unexpectedly!")

    def test_validate_bodyweight_note_text_with_invalid_input(self):
        bw_note_text = '(70.5, 71.2, 70.8, 70.6'
        with self.assertRaises(ValueError):
            validate_bodyweight_note_text(bw_note_text)


if __name__ == '__main__':
    unittest.main()
